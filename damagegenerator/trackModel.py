# ############################################################################
# 
# This software is made freely available in accordance with the simplifed BSD
# license:
# 
# Copyright (c) <2018>, <Stephen McMahon>
# All rights reserved
# Redistribution and use in source and binary forms, with or without 
# modification, are permitted provided that the following conditions are met:
#
# 1. Redistributions of source code must retain the above copyright notice, 
# this list of conditions and the following disclaimer.
# 2. Redistributions in binary form must reproduce the above copyright notice,
# this list of conditions and the following disclaimer in the documentation 
# and/or other materials provided with the distribution.
#
# THIS SOFTWARE IS PROVIDED BY THE AUTHOR AND CONTRIBUTORS ``AS IS'' AND ANY 
# EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED 
# WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE 
# DISCLAIMED. IN NO EVENT SHALL THE AUTHOR OR CONTRIBUTORS BE LIABLE FOR ANY
# DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES 
# (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; 
# LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND 
# ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT 
# (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF 
# THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.
#
# Contacts: Stephen McMahon,	stephen.mcmahon@qub.ac.uk
# 
# ############################################################################
import os
import openpyxl as pyxl

LETs = []
cumuEnergyData = []

# Read cumulative radial dose data from file
def readCumuDoseFile(filename):
	global LETs
	global cumuEnergyData

	# Use this file's path to locate input files
	filePath = os.path.realpath(__file__)
	baseDir,thisFile = os.path.split(filePath)
	wb = pyxl.load_workbook(baseDir+"/"+filename,read_only = True,data_only=True)
	sheets= wb.sheetnames
	ws = wb['Radial Energy - Cumulative']
	rows= ws.iter_rows()
	energies = [r.value for r in next(rows)][1:]
	LETList = [r.value for r in next(rows)][1:]

	cEnergyData = [[] for n in range(len(LETList))]

	for row in rows:
		r = row[0].value
		for n,c in enumerate(row[1:]):
			cEnergyData[n].append([r,c.value])

	LETs = LETList
	cumuEnergyData = cEnergyData

# Build cumulative curve for target LET, by interpolating nearest dataset
def buildCumCurve(targetLET):
	# If outside of range, just use limiting value
	if targetLET>=max(LETs): return cumuEnergyData[0]
	if targetLET<=min(LETs): return cumuEnergyData[-1]

	# Find cloest bin below desired LET, and work out fraction through bin
	lowerBin = 0
	while LETs[lowerBin+1] > targetLET: lowerBin+=1
	fractionHigh = (targetLET-LETs[lowerBin+1])/(LETs[lowerBin]-LETs[lowerBin+1])

	# Interpolate new distribution
	newCEnergy = []
	for high,low in zip(cumuEnergyData[lowerBin],cumuEnergyData[lowerBin+1]):
		newCEnergy.append([high[0],high[1]*fractionHigh+(1-fractionHigh)*low[1] ])

	return newCEnergy

# Sample radial position associated with random sample
def sampleRadialPos(randVal,radialEnergyData):
	if randVal<radialEnergyData[0][1]: return radialEnergyData[0][0]
	if randVal>radialEnergyData[-1][1]: return radialEnergyData[-1][0]

	# Iterate through radial distribution nearest bin is found
	lowerBin = 0
	while radialEnergyData[lowerBin+1][1]<randVal:lowerBin+=1

	# Interpolate to find best match
	fractionLow = ( (randVal-radialEnergyData[lowerBin][1])/ 
				    (radialEnergyData[lowerBin+1][1]-radialEnergyData[lowerBin][1]) )
	newR = radialEnergyData[lowerBin][0]*fractionLow+radialEnergyData[lowerBin+1][0]*(1-fractionLow)
	return newR